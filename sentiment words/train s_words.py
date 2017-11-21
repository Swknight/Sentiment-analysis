#-*-coding:utf-8-*-
from openpyxl import load_workbook
#from openpyxl import Workbook
wb = load_workbook(filename="train set.xlsx")
sheets = wb.get_sheet_names()
sheet_first = sheets[0]
ws = wb.get_sheet_by_name(sheet_first)
s_p_words = []
s_n_words = []
s_m_words = []
for row in range(9859,ws.max_row):
    line0 = ws.cell(row=row,column=4).value
    line1 = ws.cell(row=row,column=5).value
    if line0 is None:
        continue
    s_word = line0.split(';')
    s_word.pop()
    s_value = line1.split(';')
    s_value.pop()
    for i in range(len(s_word)):
        if int(s_value[i]) == 1:
            if s_word not in s_p_words:
             s_p_words.append(s_word[i])
        elif int(s_value[i]) == -1:
            if s_word not in s_n_words:
                s_n_words.append(s_word[i])
        else:
            if s_word not in s_m_words:
             s_m_words.append(s_word[i])
# #wb1 = Workbook()
#sheet = wb1.active
s_p_words = set(s_p_words)
s_n_words = set(s_n_words)
s_m_words = set(s_m_words)
fout0 = open("s_p_words.txt",'w')
for i in s_p_words:
#   sheet["A%d" % (i + 1)].value = s_words[i]
    fout0.write(i.strip().encode("utf-8") + "\n")
fout0.close()
#wb1.save("s_words.xlsx")
fout1 = open("s_n_words.txt",'w')
for i in s_n_words:
    fout1.write(i.strip().encode("utf-8") + "\n")
fout1.close()

fout2 = open("s_m_words.txt",'w')
for i in s_m_words:
    fout2.write(i.strip().encode("utf-8") + "\n")
fout2.close()
